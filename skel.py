#! /usr/bin/env python
"""
Create Models Skeleton from XLS

Usage:
  skel <filename> <cols> [<format>]
  skel -d <filename>
  skel -h | --help
  skel --version

Options:
  -h --help        Show this screen.
  --version        Show version.
  -d --dimensions  Print spreadsheet dimensions
  <cols>           Eg. A2:AK2
  <format>         'mapping' for bulkimport mapping style
                   'html' for django template definition list format
                   defaults to model.py style

"""
from openpyxl import load_workbook
from docopt import docopt
#from openpyxl.cell import get_column_letter
import string


def print_cols(filename, cols, format):
    """
    Parse spreadsheet file and extract column headers
    """
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
#    dimensions = sheet.calculate_dimension()
    data = sheet.range(cols)

    high_row = sheet.get_highest_row()
    for cell in data[0]:
        max_len = 0
        for row in sheet.range(cell.address + ":" + cell.address[:-1] +
                str(high_row)):
            try:
                if row[0].value and len(row[0].value) > max_len:
                    max_len = len(row[0].value)
            except TypeError:
                pass

        original_name = cell.value
        py_name = cell.value.replace(' ', '_').replace('/', '_').replace('?',
                '').replace('.', '').replace('&', '').replace('-', '').lower()
        py_name = string.strip(py_name, '_')

        if format == 'mapping':
            print "'%s': '%s'," % (original_name, py_name)
        elif format == 'html':
            print """{%% if object.%(prop)s %%}
            <dt>%(title)s</dt>
            <dd>{{ object.%(prop)s }}</dd>
{%% endif %%}

            """ % {'title': original_name, 'prop': py_name}
        else:
            print "    %s = models.CharField(\"%s\", max_length=%d, blank=True)" % (py_name, original_name.lower(), max_len + 2)


def print_dimensions(filename):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    print sheet.calculate_dimension()


if __name__ == '__main__':
    arguments = docopt(__doc__, version="Models Skeleton 0.1")

    if '<filename>' in arguments:
        if arguments['--dimensions']:
            print_dimensions(arguments['<filename>'])
        else:
            print_cols(arguments['<filename>'], arguments['<cols>'], arguments['<format>'])

